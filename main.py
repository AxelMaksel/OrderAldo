import openpyxl as op
import glob

xls_files = glob.glob('*.xls*')

def out_xls(sku_, name_, number_, price_):
    sheet2.cell(row=row_, column=1, value=sku_)
    sheet2.cell(row=row_, column=2, value=name_)
    sheet2.cell(row=row_, column=3, value=number_)
    sheet2.cell(row=row_, column=4, value=price_)


filename = 'Заказ клиента1.xlsx'
subcategories_dict = {}

wb = op.load_workbook(filename, data_only=True)
sheet = wb.active
out_xlsx = op.Workbook()
max_rows = sheet.max_row
sheet2 = out_xlsx.active
row_ = 1
# list[]=[]
while xls_files:
    name = xls_files.pop()
    if name == "aldo_1c.xlsx":
        continue

    for i in range(13, max_rows + 1):
        sku = sheet.cell(row=i, column=6).value  # артикул товара
        name = sheet.cell(row=i, column=12).value  # наименование товара
        number = sheet.cell(row=i, column=38).value  # кол-во товара
        price = sheet.cell(row=i, column=48).value  # стоимость товара

        if not sku:
            continue

        print(sku, name, number, price)
        out_xls(sku, name, number, price)
        row_ += 1

out_xlsx.save('aldo_1c.xlsx')
